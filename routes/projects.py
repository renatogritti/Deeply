from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file, g
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import json
from openpyxl import Workbook, load_workbook
import os
from werkzeug.utils import secure_filename
import tempfile
import uuid  # Adicionar esta importação no topo do arquivo

from app import *
from models.db import *

def init_app(app):

    """
    KCRUD: Kanban User CRUD Operations
    This module provides REST API endpoints for managing Kanban users.
    """
    @app.route('/api/projects', methods=['GET'])
    def get_projects():
        projects = Project.query.order_by(Project.name).all()
        return jsonify([project.to_dict() for project in projects])

    @app.route('/api/projects', methods=['POST'])
    def create_project():
        """Create a new project with phases."""
        try:
            data = request.json
            
            # Validações
            if not data.get('name'):
                return jsonify({"success": False, "error": "Project name is required"}), 400
            
            if not data.get('phases') or len(data['phases']) == 0:
                return jsonify({"success": False, "error": "At least one phase is required"}), 400
                
            project = Project(
                name=data['name'],
                description=data.get('description', '')
            )
            
            # Adicionar fases do projeto
            for idx, phase_data in enumerate(data['phases']):
                if not phase_data.get('name'):
                    return jsonify({"success": False, "error": f"Phase {idx+1} name is required"}), 400
                    
                phase = Phase(
                    name=phase_data['name'],
                    order=idx,
                    project=project
                )
                db.session.add(phase)
                
            db.session.add(project)
            db.session.commit()
            return jsonify({"success": True, "project": project.to_dict()})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 400

    @app.route('/api/projects/<int:project_id>', methods=['PUT'])
    def update_project(project_id):
        """Update project and its phases."""
        try:
            project = Project.query.get_or_404(project_id)
            data = request.json
            
            # Validações
            if 'name' in data and not data['name']:
                return jsonify({"success": False, "error": "Project name cannot be empty"}), 400
                
            if 'phases' in data and len(data['phases']) == 0:
                return jsonify({"success": False, "error": "At least one phase is required"}), 400
            
            # Atualizar dados básicos do projeto
            if 'name' in data:
                project.name = data['name']
            if 'description' in data:
                project.description = data['description']
                
            # Atualizar fases
            if 'phases' in data:
                # Primeiro, verifique se há cards em alguma fase
                cards_count = KanbanCard.query.filter_by(project_id=project_id).count()
                if cards_count > 0:
                    return jsonify({
                        "success": False, 
                        "error": "Cannot modify phases while there are cards in the project"
                    }), 400
                
                # Se não houver cards, pode atualizar as fases
                # Remover fases antigas
                for phase in project.phases:
                    db.session.delete(phase)
                
                # Adicionar novas fases
                for idx, phase_data in enumerate(data['phases']):
                    if not phase_data.get('name'):
                        return jsonify({"success": False, "error": f"Phase {idx+1} name is required"}), 400
                        
                    phase = Phase(
                        name=phase_data['name'],
                        order=idx,
                        project=project
                    )
                    db.session.add(phase)
            
            db.session.commit()
            return jsonify({"success": True, "project": project.to_dict()})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 400

    @app.route('/api/projects/<int:project_id>/phases', methods=['GET'])
    def get_project_phases(project_id):
        try:
            project = Project.query.get_or_404(project_id)
            return jsonify([phase.to_dict() for phase in project.phases])
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 404

    @app.route('/api/projects/<int:project_id>', methods=['DELETE'])
    def delete_project(project_id):
        try:
            project = Project.query.get_or_404(project_id)
            db.session.delete(project)
            db.session.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 400

    # Atualize ou adicione esta rota
    @app.route('/api/projects/<int:project_id>', methods=['GET'])
    def get_project(project_id):
        """Get a specific project by ID."""
        try:
            project = Project.query.get_or_404(project_id)
            return jsonify(project.to_dict())
        except Exception as e:
            return jsonify({"error": str(e)}), 404

    @app.route('/api/cards/phases/<int:project_id>', methods=['GET'])
    def get_phases_for_card(project_id):
        """Get phases for a specific project for card creation/editing."""
        try:
            project = Project.query.get_or_404(project_id)
            phases = [phase.to_dict() for phase in project.phases]
            return jsonify({"success": True, "phases": phases})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 404

    @app.route('/api/projects/<int:project_id>/export', methods=['GET'])
    def export_project(project_id):
        try:
            project = Project.query.get_or_404(project_id)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Project Export"
            
            # Headers atualizados
            headers = [
                'Phase', 'Card Title', 'Description', 'Time Estimate', 
                'Start Date', 'Due Date', 'Users', 'Tags', 
                'Comments', 'Percentage'
            ]
            ws.append(headers)
            
            for phase in project.phases:
                for card in phase.cards:
                    users = ";".join([user.email for user in card.users]) if card.users else ''
                    tags = ";".join([tag.name for tag in card.tags]) if card.tags else ''
                    
                    # Formatação das datas
                    start_date = card.start_date.strftime('%Y-%m-%d') if card.start_date else ''
                    due_date = card.deadline.strftime('%Y-%m-%d') if card.deadline else ''
                    
                    # Formatação dos comentários
                    comments = "; ".join([f"{c.user.email}: {c.content}" for c in card.comments]) if card.comments else ''
                    
                    row = [
                        phase.name,
                        card.title,
                        card.description,
                        card.tempo,
                        start_date,
                        due_date,
                        users,
                        tags,
                        comments,
                        card.percentage
                    ]
                    ws.append(row)
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Limita a largura máxima
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f"project_export_{project_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            wb.save(filepath)
            
            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 400

    @app.route('/api/projects/<int:project_id>/import', methods=['POST'])
    def import_project(project_id):
        temp_log = None
        try:
            if 'file' not in request.files:
                return jsonify({"success": False, "error": "No file uploaded"}), 400
                
            file = request.files['file']
            if not file.filename.endswith('.xlsx'):
                return jsonify({"success": False, "error": "Invalid file format"}), 400

            project = Project.query.get_or_404(project_id)
            error_log = []
            temp_log = tempfile.NamedTemporaryFile(delete=False, mode='w', suffix='.txt')
            
            def parse_excel_date(date_value):
                """Converte uma data do Excel para datetime"""
                if not date_value:
                    return None
                    
                if isinstance(date_value, datetime):
                    return date_value
                    
                try:
                    # Tenta converter string para data
                    if isinstance(date_value, str):
                        # Tenta diferentes formatos de data
                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                            try:
                                return datetime.strptime(date_value.strip(), fmt)
                            except ValueError:
                                continue
                        
                    # Se for um número do Excel (número de dias desde 1900)
                    if isinstance(date_value, (int, float)):
                        return datetime(1899, 12, 30) + timedelta(days=int(date_value))
                        
                except Exception as e:
                    return None
                    
                return None
            
            try:
                wb = load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                
                with db.session.no_autoflush:
                    for row in ws.iter_rows(min_row=2):
                        try:
                            row_data = dict(zip(headers, [cell.value for cell in row]))
                            if not row_data['Card Title']:
                                continue
                            
                            # Validar fase
                            phase = next((p for p in project.phases if p.name == row_data['Phase']), None)
                            if not phase:
                                raise ValueError(f"Phase not found: {row_data['Phase']}")
                            
                            card_title = row_data['Card Title'].strip()
                            existing_card = KanbanCard.query.filter_by(
                                project_id=project_id,
                                title=card_title
                            ).first()

                            # Processamento das datas usando a nova função
                            start_date = parse_excel_date(row_data.get('Start Date'))
                            if not start_date and row_data.get('Start Date'):
                                error_log.append(f"Warning: Invalid start date format for card '{card_title}': {row_data['Start Date']}")

                            due_date = parse_excel_date(row_data.get('Due Date'))
                            if not due_date and row_data.get('Due Date'):
                                error_log.append(f"Warning: Invalid due date format for card '{card_title}': {row_data['Due Date']}")

                            # Processamento de usuários
                            users = []
                            if row_data['Users']:
                                for email in row_data['Users'].split(';'):
                                    user = Team.query.filter_by(email=email.strip()).first()
                                    if not user:
                                        error_log.append(f"Warning: User not found: {email}")
                                    else:
                                        users.append(user)
                                        
                            # Processamento de tags
                            tags = []
                            if row_data['Tags']:
                                for tag_name in row_data['Tags'].split(';'):
                                    tag = Tag.query.filter_by(name=tag_name.strip()).first()
                                    if not tag:
                                        error_log.append(f"Warning: Tag not found: {tag_name}")
                                    else:
                                        tags.append(tag)

                            # Processamento de comentários
                            comments = []
                            if row_data['Comments']:
                                for comment_text in row_data['Comments'].split(';'):
                                    parts = comment_text.split(':', 1)
                                    if len(parts) == 2:
                                        user_email, content = parts
                                        user = Team.query.filter_by(email=user_email.strip()).first()
                                        if user:
                                            comment = Comment(
                                                content=content.strip(),
                                                user=user
                                            )
                                            comments.append(comment)

                            # Atualizar ou criar card com as novas datas
                            if existing_card:
                                existing_card.description = row_data['Description']
                                existing_card.tempo = row_data['Time Estimate']
                                existing_card.start_date = start_date
                                existing_card.deadline = due_date
                                existing_card.phase_id = phase.id
                                existing_card.users = users
                                existing_card.tags = tags
                                existing_card.percentage = row_data.get('Percentage', 0)
                                
                                # Atualizar comentários
                                for comment in comments:
                                    comment.card = existing_card
                                    db.session.add(comment)
                            else:
                                new_card = KanbanCard(
                                    id=str(uuid.uuid4()),
                                    project_id=project_id,
                                    phase_id=phase.id,
                                    title=card_title,
                                    description=row_data['Description'],
                                    tempo=row_data['Time Estimate'],
                                    start_date=start_date,
                                    deadline=due_date,
                                    percentage=row_data.get('Percentage', 0),
                                    users=users,
                                    tags=tags
                                )
                                db.session.add(new_card)
                                
                                # Adicionar comentários ao novo card
                                for comment in comments:
                                    comment.card = new_card
                                    db.session.add(comment)
                            
                            db.session.flush()
                                
                        except Exception as e:
                            error_log.append(f"Error in row {row[0].row}: {str(e)}")

                    if error_log:
                        # Escrever logs mas continuar com a importação
                        for error in error_log:
                            temp_log.write(error + '\n')
                        temp_log.close()
                        
                    db.session.commit()
                    return jsonify({
                        "success": True,
                        "has_warnings": bool(error_log),
                        "warning_count": len(error_log)
                    })
                    
            finally:
                wb.close()
                if temp_log and not temp_log.closed:
                    temp_log.close()
                if error_log:
                    session['error_log_path'] = temp_log.name
                
        except Exception as e:
            if temp_log and not temp_log.closed:
                temp_log.close()
            return jsonify({"success": False, "error": str(e)}), 400

    @app.route('/api/projects/<int:project_id>/error_log')
    def download_error_log(project_id):
        if 'error_log_path' not in session:
            return jsonify({"error": "No error log found"}), 404
            
        error_file = session['error_log_path']
        try:
            # Enviar o arquivo
            response = send_file(
                error_file,
                as_attachment=True,
                download_name=f"import_errors_{project_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            )
            
            # Adicionar callback para remover o arquivo após o envio
            @response.call_on_close
            def remove_file():
                try:
                    if os.path.exists(error_file):
                        os.unlink(error_file)
                except Exception as e:
                    print(f"Error removing temp file: {e}")
                finally:
                    session.pop('error_log_path', None)
                    
            return response
            
        except Exception as e:
            session.pop('error_log_path', None)
            return jsonify({"error": str(e)}), 500
